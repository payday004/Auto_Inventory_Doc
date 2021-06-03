import tkinter as tk
import openpyxl
import os
import mailmerge

fields = ('First Name', 'Last Name', 'Date', 'Asset', 'Serial Number', 'Asset List')
password = "examplepassword"
domain = "highstretit.com"


# created pop up message for GUI
def popup_msg(msg):

    popup = tk.Tk()
    popup.wm_title("warning!")
    label = tk.Label(popup, text=msg)
    label.pack(side="top", fill="x", padx=10, pady=10)
    okay_button = tk.Button(popup, text="Okay", command=popup.destroy)
    okay_button.pack(padx=10, pady=10)
    popup.mainloop()


# resets GUI entries after successful update
def clear_entries(entries):

    entries['First Name'].delete(0, tk.END)
    entries['First Name'].insert(0, "0")

    entries['Last Name'].delete(0, tk.END)
    entries['Last Name'].insert(0, "0")

    entries['Date'].delete(0, tk.END)
    entries['Date'].insert(0, "0")

    entries['Asset'].delete(0, tk.END)
    entries['Asset'].insert(0, "0")

    entries['Serial Number'].delete(0, tk.END)
    entries['Serial Number'].insert(0, "0")

    entries['Asset List'].delete(0, tk.END)
    entries['Asset List'].insert(0, "0")


# linear search for excel "Database"
def search_excel_sheet(serial_number, wb):
    ws = wb.active

    # search excel file for inventory id
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        if str(row[1].value) == serial_number:
            return row


# create tkinter window to input data
def make_form(root, in_fields):

    entries = {}
    for field in in_fields:
        # print(field)
        row = tk.Frame(root)
        lab = tk.Label(row, width=22,
                       text=field+": ",
                       anchor='w')
        ent = tk.Entry(row)
        ent.insert(0, "0")
        row.pack(side=tk.TOP,
                 fill=tk.X,
                 padx=5,
                 pady=5)
        lab.pack(side=tk.LEFT)
        ent.pack(side=tk.RIGHT,
                 expand=tk.YES,
                 fill=tk.X)
        entries[field] = ent

    return dict(entries)


# main method for CMD prompt
def run_cmd_prompt():

    cmd_first_name = input("enter first name:    \t")
    cmd_last_name = input("enter last name:     \t")
    cmd_date = input("enter date:          \t")
    cmd_asset = input("enter asset:         \t")
    cmd_serial_number = input("enter serial number: \t")
    cmd_asset_list = input("enter asset list:    \t")

    # set up data.xlsx
    input_file = str(os.getcwd()) + '\\' + "data.xlsx"
    wb = openpyxl.load_workbook(filename=input_file)

    row = search_excel_sheet(cmd_serial_number, wb)

    if row is not None:
        print("\n" + "found " + str(row[3].value) + "...")
        row[0].value = str(cmd_last_name + ", " + cmd_first_name)

        wb.save('data.xlsx')
        print("\n" + "excel sheet updated...")

        # set up form.docx for output
        template_hard = str(os.getcwd()) + '\\' + "form_hard.docx"
        template_pass = str(os.getcwd()) + '\\' + "form_pass.docx"

        # create and save documents
        # hardware doc
        save_string_1 = cmd_last_name + "_hardware" + ".docx"
        save_string_2 = cmd_last_name + "_password" + ".docx"

        document_1 = mailmerge.MailMerge(template_hard)
        document_1.merge(NAME=cmd_first_name + ' ' + cmd_last_name,
                         DATE=cmd_date,
                         ASSET=cmd_asset,
                         SERIAL_NUM=cmd_serial_number,
                         LIST=cmd_asset_list)
        document_1.write("output" + '\\' + save_string_1)

        # password doc
        document_2 = mailmerge.MailMerge(template_pass)
        username = cmd_first_name + '.' + cmd_last_name
        document_2.merge(PASSWORD=password,
                         USERNAME=username,
                         DOMAIN=domain,
                         EMAIL_ADDRESS=username + '@' + domain)
        document_2.write("output" + '\\' + save_string_2)

        print("\n" + "output documents created..." + "\n")

        continue_bool = int(input("1 to add another entry, 2 to exit: "))

        if continue_bool == 1:
            run_cmd_prompt()

    else:
        print("\n" + "Error: that serial number was not found, try a valid serial number" + "\n")
        run_cmd_prompt()


# main method for GUI prompt
def run_enter(entries):
    # get information
    first_name = entries['First Name'].get()
    last_name = entries['Last Name'].get()
    date = entries['Date'].get()
    asset = entries['Asset'].get()
    serial_number = str(entries['Serial Number'].get())
    asset_list = entries['Asset List'].get()

    # set up data.xlsx
    input_file = str(os.getcwd()) + '\\' + "data.xlsx"
    wb = openpyxl.load_workbook(filename=input_file)

    row = search_excel_sheet(serial_number, wb)

    if row is not None:

        print("found " + str(row[3].value))
        row[0].value = str(last_name + ", " + first_name)

        # update information
        wb.save('data.xlsx')

        # set up form.docx for output
        template_hard = str(os.getcwd()) + '\\' + "form_hard.docx"
        template_pass = str(os.getcwd()) + '\\' + "form_pass.docx"

        # create and save documents
        # hardware doc
        save_string_1 = last_name + "_hardware" + ".docx"
        save_string_2 = last_name + "_password" + ".docx"

        document_1 = mailmerge.MailMerge(template_hard)
        document_1.merge(NAME=first_name + ' ' + last_name,
                         DATE=date,
                         ASSET=asset,
                         SERIAL_NUM=serial_number,
                         LIST=asset_list)
        document_1.write("output" + '\\' + save_string_1)

        # password doc
        document_2 = mailmerge.MailMerge(template_pass)
        username = first_name + '.' + last_name
        document_2.merge(PASSWORD=password,
                         USERNAME=username,
                         DOMAIN=domain,
                         EMAIL_ADDRESS=username + '@' + domain)
        document_2.write("output" + '\\' + save_string_2)

        clear_entries(entries)
        popup_msg("sheet updated and documents created")

    else:
        popup_msg("the serial number was not found")


if __name__ == '__main__':

    gui_select = int(input("1 for GUI, 2 for CMD:\t "))

    if gui_select == 1:

        win = tk.Tk()
        ent_form = make_form(win, fields)

        b1 = tk.Button(win, text='Enter', command=(lambda e=ent_form: run_enter(e)))
        b1.pack(side=tk.LEFT, padx=5, pady=5)

        b2 = tk.Button(win, text='Quit', command=win.quit)
        b2.pack(side=tk.RIGHT, padx=5, pady=5)

        win.mainloop()

    else:

        run_cmd_prompt()


# 123456
# 123451
# 123446
